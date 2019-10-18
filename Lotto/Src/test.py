#!/usr/bin/env python
# -*- coding:utf-8 -*-
# Author:Charioty
# time:2019/9/3

import requests
from bs4 import BeautifulSoup
import openpyxl
import time
import random

def get_headers():
    '''
    随机获取一个headers
    '''
    user_agents =  ['Mozilla/5.0(Macintosh;U;IntelMacOSX10_6_8;en-us)AppleWebKit/534.50(KHTML,likeGecko)Version/5.1Safari/534.50','Mozilla/5.0(Windows;U;WindowsNT6.1;en-us)AppleWebKit/534.50(KHTML,likeGecko)Version/5.1Safari/534.50','Mozilla/5.0(compatible;MSIE9.0;WindowsNT6.1;Trident/5.0','Mozilla/4.0(compatible;MSIE8.0;WindowsNT6.0;Trident/4.0)','Mozilla/5.0(Macintosh;IntelMacOSX10.6;rv:2.0.1)Gecko/20100101Firefox/4.0.1','Mozilla/5.0(WindowsNT6.1;rv:2.0.1)Gecko/20100101Firefox/4.0.1','Mozilla/4.0(compatible;MSIE7.0;WindowsNT5.1;Trident/4.0;SE2.XMetaSr1.0;SE2.XMetaSr1.0;.NETCLR2.0.50727;SE2.XMetaSr1.0)','Mozilla/4.0(compatible;MSIE7.0;WindowsNT5.1;360SE)']
    headers = {'User-Agent':random.choice(user_agents)}
    return headers

def get_all_Urls():
    """
        获取所有所需链接
    """
    url = 'http://kaijiang.500.com/ssq.shtml?0_ala_baidu'
    book_list = []
    r = requests.get(url, headers = get_headers(), timeout=30)
    r.encoding = 'GBK'
    soup = BeautifulSoup(r.text, 'lxml')
    #整理数据，找出所需的链接都有哪些，写入数组。
    book_ul = soup.find_all('div', {'class': 'iSelectList'})
    book_ps = book_ul[0].find_all('a')
    for book_p in book_ps:
        book_url = book_p.get('href')
        book_list.append(book_url)
    return book_list

def Recom(args):
    Num_R1 = []
    Num_R = []
    if args:
        for arg in args:
            if arg == ' ' or arg == '|' or arg == '\t' or arg == '\r' or arg == '\n':
                continue
            else:
                Num_R1.append(arg)
        if Num_R1:
            i = 0
            while i <12:
                Num_R.append(Num_R1[i]+Num_R1[i+1])
                i = i+2
            return Num_R
        else:
            return Num_R
    else:
        return Num_R

if __name__ == '__main__':
    print('数据正在收集，请耐心等待')
    Urls = get_all_Urls()
    # random.shuffle(Urls)
    book = openpyxl.Workbook()
    sheet1 = book.active
    sheet1.title = 'data'
    sheet1["A1"]='Number'  #第0行第0列
    sheet1.merge_cells('B1:H1')
    sheet1["B1"]='orderball'
    sheet1.merge_cells('I1:O1')
    sheet1["I1"] = 'orgball'
    #print(len(Urls))#1872
    for url in Urls:
        row = sheet1.max_row
        time_s = 0
        charNum=[]
        time.sleep(time_s)
        m = requests.get(url, headers = get_headers(),timeout=30)
        m.encoding = 'GBK'
        htmltext = BeautifulSoup(m.text, 'lxml')
        #查找期号
        Num_ul = htmltext.find_all('td',{'class': 'td_title01'})
        Num_div = Num_ul[0].find_all('strong')
        if Num_div[0].string == '19102':
            charNum.append(Num_div[0].string)
            #查找排列数
            Lnum_ul = htmltext.find_all('table',{'width': '100%','border': '0','cellspacing': '0','cellpadding': '1'})
            Lnum_reds = Lnum_ul[0].find_all('li',{'class': 'ball_red'})
            for Lnum_red in Lnum_reds:
                charNum.append(Lnum_red.string)
            Lnum_blues = Lnum_ul[0].find_all('li',{'class': 'ball_blue'})
            for Lnum_blue in Lnum_blues:
                charNum.append(Lnum_blue.string)
            #查找不按照顺序排列的数
            all_numurl = Lnum_ul[0].find_all('td')[3].string
            print(all_numurl)
            for allnum in Recom(all_numurl):
                charNum.append(allnum)
            charNum = list(map(int, charNum))#转换类型，因为append只吃list的int类型
            sheet1.append(charNum)
            print('已完成到第'+str(row)+'条数据')
            book.save('Old_data.xlsx')
        else:
            continue
    book.close()
    print('数据收集已完成，请查阅')
