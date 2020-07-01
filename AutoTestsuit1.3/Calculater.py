#!/usr/bin/env python
# -*- coding:utf-8 -*-
# Author:Charioty
# time:2020/2/21

import sqlite3
import configparser
import openpyxl
import threading
from openpyxl.styles import Alignment, Border, Side

#读取Json文件内容
def readJsonFileToStr(file_name):
    with open(file_name, 'r', encoding='utf-8') as f:
        text = f.readlines()
        f.close()
    return text
#获得配置文件
def config():
    config = configparser.ConfigParser()
    config.read("option.ini")
    Level1 = config.get("FAR","Level1")
    Level2 = config.get("FAR","Level2")
    Pre_v = config.get("option","pre_v")
    Match_v = config.get("option","match_v")
    Image_s = config.get("option","image_s")
    Case_num = config.get("option","case_num")
    return Level1,Level2,Pre_v,Match_v,Image_s,Case_num

#获得学习阈值
def Learn_config(url):
    Url = url+'\SensorData.json'
    text = readJsonFileToStr(Url)
    for i in text:
        if 'isLearningOnLine' in i :
            LS=i[i.find(":")+1:-2].strip()
        if '"clt":'in i :
            LT = i[i.find(":") + 1:-2].strip()
        if '"dclt":'in i :
            DLT = i[i.find(":") + 1:-2].strip()
        if '"slt":'in i :
            SLT = i[i.find(":") + 1:-2].strip()
        if '"dslt":'in i :
            DSLT = i[i.find(":") + 1:-1].strip()
    return LS,LT,SLT,DLT,DSLT

#根据要求查询相应阈值
def search_THS(L1,Url,QUERY_THR_0):
    conn = sqlite3.connect(Url+'.\MatchResult.dat')
    c = conn.cursor()
    Data_Sum_nor= c.execute("SELECT count(*) FROM MatchResult WHERE MatchType = 0 and SampleID not like '%light%' and SampleID like '%FT$0%' ").fetchall()
    num1 = round(Data_Sum_nor[0][0] / int(L1))+1
    THS_temp_nor = c.execute("select Score from" +QUERY_THR_0+ "where MatchType = 0 and SampleID not like '%light%' and SampleID  like '%FT$0%' order by Score DESC limit ?", (num1,)).fetchall()
    Data_Sum_dry= c.execute("SELECT count(*) FROM MatchResult WHERE MatchType = 0 and SampleID not like '%light%' and SampleID like '%FT$1%' ").fetchall()
    num2 = round(Data_Sum_dry[0][0] / int(L1))+1
    THS_temp_dry = c.execute("select Score from" +QUERY_THR_0+ "where MatchType = 0 and SampleID not like '%light%' and SampleID like '%FT$1%' order by Score DESC limit ?", (num2,)).fetchall()
    return THS_temp_nor[num1-1][0]+1,THS_temp_dry[num2-1][0]+1#查出分数+1
    c.close()
    conn.close()

#获得场景集合
def get_scene(Url):
    scene = []
    conn = sqlite3.connect(Url+'.\MatchResult.dat')
    c = conn.cursor()
    Namp = c.execute("SELECT SampleId from MatchResult where SampleId like 'U%'").fetchall()
    Samp = c.execute("SELECT SampleId from MatchResult where SampleId like 'Z%'").fetchall()
    if Samp:
        for Sample in Samp:
            scene.append(Sample[0][Sample[0].find('_')+8:])
    scene = list(set(scene))
    scene.sort()
    if Namp:
        scene.insert(0,'normal')
    return scene

#计算FRR
def calculate_FRR(Lx,scene,Url,QUERY_THR_1):
    conn = sqlite3.connect(Url+'.\MatchResult.dat')
    c = conn.cursor()
    if scene == 'normal':
        X ='U%'
        sum_normal = c.execute("SELECT count(*) from " +QUERY_THR_1+ " where MatchType = 1 and SampleId like ?", (X,)).fetchall();
        num_normal = c.execute("SELECT count(*) from " +QUERY_THR_1+ " where MatchType = 1 and SampleId like ? and Score < ?", (X,Lx)).fetchall();
        FRR = "%.2f%%" % (num_normal[0][0]/sum_normal[0][0]*100)
    else:
        X = '%'+scene
        sum_normal = c.execute("SELECT count(*) from " +QUERY_THR_1+ " where MatchType = 1 and SampleId like ?", (X,)).fetchall();
        num_normal = c.execute("SELECT count(*) from " +QUERY_THR_1+ " where MatchType = 1 and SampleId like ? and Score < ?", (X,Lx)).fetchall();
        FRR = "%.2f%%" % (num_normal[0][0]/sum_normal[0][0]*100)
    return FRR
#以下函数为填充对应表格
def learn_insert_scene(sheet1,scene):
    for i,j in enumerate(scene):
        sheet1['C'+str(i+2)] = j
def learn_insert_L1(sheet1, Levelnor1,Leveldry1,scene):
    for i,j in enumerate(scene):
        if 'dry' in j:
            sheet1['D' + str(i + 2)] = Leveldry1
        else:
            sheet1['D' + str(i + 2)] = Levelnor1
def learn_insert_L2(sheet1, Levelnor1,Leveldry1,scene,Url,QUERY_THR_1):
    for i,j in enumerate(scene):
        if  'dry' in j:
            sheet1['E' + str(i + 2)] = calculate_FRR(int(Leveldry1), j,Url,QUERY_THR_1)
        else:
            sheet1['E' + str(i + 2)] = calculate_FRR(int(Levelnor1), j,Url,QUERY_THR_1)
def learn_insert_L3(sheet1, Levelnor2,Leveldry2,scene):
    for i,j in enumerate(scene):
        if 'dry' in j:
            sheet1['F'+str(i+2)] = Leveldry2
        else:
            sheet1['F'+str(i+2)] = Levelnor2
def learn_insert_L4(sheet1, Levelnor2,Leveldry2,scene,Url,QUERY_THR_1):
    for i,j in enumerate(scene):
        if 'dry' in j:
            sheet1['G'+str(i+2)] = calculate_FRR(int(Leveldry2), j,Url,QUERY_THR_1)
        else:
            sheet1['G'+str(i+2)] = calculate_FRR(int(Levelnor2), j,Url,QUERY_THR_1)

def no_insert_scene(sheet1,scene,num):
    for i,j in enumerate(scene):
        sheet1['C'+str(i+num+2)] = j
def no_insert_L1(sheet1,Levelnor1,Leveldry1,scene,num):
    for i,j in enumerate(scene):
        if 'dry' in j:
            sheet1['D' + str(i +num+ 2)] = Leveldry1
        else:
            sheet1['D' + str(i +num+ 2)] = Levelnor1
def no_insert_L2(sheet1, Levelnor1,Leveldry1,scene,Url,num,QUERY_THR_1):
    for i,j in enumerate(scene):
        if  'dry' in j:
            sheet1['E' + str(i +num+ 2)] = calculate_FRR(int(Leveldry1), j,Url,QUERY_THR_1)
        else:
            sheet1['E' + str(i +num+ 2)] = calculate_FRR(int(Levelnor1), j,Url,QUERY_THR_1)
def no_insert_L3(sheet1,Levelnor2,Leveldry2,scene,num):
    for i,j in enumerate(scene):
        if 'dry' in j:
            sheet1['F'+str(i+num+2)] = Leveldry2
        else:
            sheet1['F'+str(i+num+2)] = Levelnor2
def no_insert_L4(sheet1, Levelnor2,Leveldry2,scene,Url,num,QUERY_THR_1):
    for i,j in enumerate(scene):
        if 'dry' in j:
            sheet1['G'+str(i+num+2)] = calculate_FRR(int(Leveldry2), j,Url,QUERY_THR_1)
        else:
            sheet1['G'+str(i+num+2)] = calculate_FRR(int(Levelnor2), j,Url,QUERY_THR_1)

def Calculater(Url):
    QUERY_THR_0 = "(select SampleId, MatchType,max(Score) as Score from MatchResult where MatchType=0 group by TemplateId , substr(SampleId,0,23)  having count(*) >= 1 )"
    QUERY_THR_1 = "(select SampleId, MatchType,max(Score) as Score from MatchResult where MatchType=1 group by substr(SampleId,0,23)  having count(*) >= 1 )"
    scene=get_scene(Url)
    num = int(len(scene))
    book = openpyxl.load_workbook('result.xlsx')
    sheet1 = book.get_sheet_by_name('FRR')
    sheet1["A1"] = 'information'
    sheet1["B1"] = 'learn'
    sheet1["C1"] = 'scene'
    sheet1["D1"] = 'threshold'+config()[0]
    sheet1["E1"] = 'FAR'+config()[0]
    sheet1["F1"] = 'threshold'+config()[1]
    sheet1["G1"] = 'FAR'+config()[1]
    sheet1.merge_cells('A2:A'+str(num*2+1))
    sheet1["A2"] = 'Preprocessing：'+config()[2]+' \n'+'match：'+config()[3]+' \n '+'image：'+config()[4]+' \n'+'total：'+config()[5]
    sheet1.merge_cells('B2:B'+str(num+1))
    sheet1.merge_cells('B'+str(num+2)+':B'+str(num*2+1))
    Level_nor_1 = search_THS(config()[0],Url,QUERY_THR_0)[0]
    Level_dry_1 = search_THS(config()[0],Url,QUERY_THR_0)[1]
    Level_nor_2 = search_THS(config()[1],Url,QUERY_THR_0)[0]
    Level_dry_2 = search_THS(config()[1],Url,QUERY_THR_0)[1]
    if Learn_config(Url)[0]== '1':
        sheet1['B2'] = 'LT:'+Learn_config(Url)[1]+' '+'SLT:'+Learn_config(Url)[2]+' '+'DLT:'+Learn_config(Url)[3]+' '+'DSLT:'+Learn_config(Url)[4]
        t1 = threading.Thread(target=learn_insert_scene,args=(sheet1,scene))
        t2 = threading.Thread(target=learn_insert_L1,args=(sheet1,Level_nor_1,Level_dry_1,scene))
        t3 = threading.Thread(target=learn_insert_L2,args=(sheet1,Level_nor_1,Level_dry_1,scene,Url,QUERY_THR_1))
        t4 = threading.Thread(target=learn_insert_L3,args=(sheet1,Level_nor_2,Level_dry_2,scene))
        t5 = threading.Thread(target=learn_insert_L4,args=(sheet1,Level_nor_2,Level_dry_2,scene,Url,QUERY_THR_1))
    else:
        sheet1['B'+str(num+2)] = 'nolearn'
        t1 = threading.Thread(target=no_insert_scene,args=(sheet1,scene,num))
        t2 = threading.Thread(target=no_insert_L1,args=(sheet1,Level_nor_1,Level_dry_1,scene,num))
        t3 = threading.Thread(target=no_insert_L2,args=(sheet1,Level_nor_1,Level_dry_1,scene,Url,num,QUERY_THR_1))
        t4 = threading.Thread(target=no_insert_L3,args=(sheet1,Level_nor_2,Level_dry_2,scene,num))
        t5 = threading.Thread(target=no_insert_L4,args=(sheet1,Level_nor_2,Level_dry_2,scene,Url,num,QUERY_THR_1))
    print('wait for calculate')
    t1.start()
    t2.start()
    t3.start()
    t4.start()
    t5.start()
    #判断线程是否全部跑完
    while True:
        if t1.isAlive()== False and t2.isAlive()== False and t3.isAlive()== False and t4.isAlive()== False and t5.isAlive()== False:
            thin = Side(border_style="thin")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            for row in sheet1['A1:G'+str(num*2+1)]:
                for cell in row:
                    cell.border = border
                    cell.alignment = alignment
            book.save('result.xlsx')
            book.close()
            print('"Mission Complete"')
            break
        else:
            continue

