#!/usr/bin/env python
# -*- coding:utf-8 -*-
# Author:Charioty
# time:2020/2/21

import sqlite3
import configparser
import openpyxl
import threading
from openpyxl.styles import Alignment, Border, Side

#获得配置文件
def config():
    config = configparser.ConfigParser()
    config.read("option.ini")
    Level1 = config.get("FAR","Level1")
    Level2 = config.get("FAR","Level2")
    Level3 = config.get("FAR","Level3")
    Level4 = config.get("FAR","Level4")
    Pre_v = config.get("option","pre_v")
    Match_v = config.get("option","match_v")
    Image_s = config.get("option","image_s")
    Case_num = config.get("option","case_num")
    return Level1,Level2,Level3,Level4,Pre_v,Match_v,Image_s,Case_num

#获得学习阈值
def Learn_config():
    learn_c = configparser.ConfigParser()
    learn_c.read("config.ini")
    LS = learn_c.get("General","LearnOnline")
    LT = learn_c.get("General","LearnThreshold")
    LVT = learn_c.get("General","LVThreshold")
    return LS,LT,LVT

#根据要求查询相应阈值
def search_THS(L1):
    conn = sqlite3.connect('MatchResult.dat')
    c = conn.cursor()
    Data_Sum = c.execute("SELECT count(*) FROM MatchResult WHERE MatchType = 0").fetchall()
    num1 = round(Data_Sum[0][0] / int(L1))+1
    THS_temp = c.execute("select Score from MatchResult where MatchType = 0 order by Score DESC limit ?", (num1,)).fetchall()
    return THS_temp[num1-1][0]+1#查出分数+1
    c.close()
    conn.close()

#获得场景集合
def get_scene():
    scene = []
    conn = sqlite3.connect('MatchResult.dat')
    c = conn.cursor()
    Namp = c.execute("SELECT SampleId from MatchResult where SampleId like 'U%'").fetchall()
    Samp = c.execute("SELECT SampleId from MatchResult where SampleId like 'Z%'").fetchall()
    if Samp:
        for Sample in Samp:
            scene.append(Sample[0][Sample[0].find('_')+8:])
    scene = list(set(scene))
    scene.sort()
    if Namp:
        scene.insert(0,'normal')
    return scene

#计算FRR
def calculate_FRR(Lx,scene):
    conn = sqlite3.connect('MatchResult.dat')
    c = conn.cursor()
    if scene == 'normal':
        X ='U%'
        sum_normal = c.execute("SELECT count(*) from MatchResult where MatchType = 1 and SampleId like ?", (X,)).fetchall();
        num_normal = c.execute("SELECT count(*) from MatchResult where MatchType = 1 and SampleId like ? and Score < ?", (X,Lx)).fetchall();
        FRR = "%.2f%%" % (num_normal[0][0]/sum_normal[0][0]*100)
    else:
        X = '%'+scene
        sum_normal = c.execute("SELECT count(*) from MatchResult where MatchType = 1 and SampleId like ?", (X,)).fetchall();
        num_normal = c.execute("SELECT count(*) from MatchResult where MatchType = 1 and SampleId like ? and Score < ?", (X,Lx)).fetchall();
        FRR = "%.2f%%" % (num_normal[0][0]/sum_normal[0][0]*100)
    return FRR
#以下函数为填充对应表格
def learn_insert_scene(sheet1,scene):
    for i,j in enumerate(scene):
        sheet1['C'+str(i+3)] = j
def learn_insert_L1(sheet1, Level1,scene):
    for i,j in enumerate(scene):
        sheet1['D'+str(i+3)] = calculate_FRR(int(Level1), j)
def learn_insert_L2(sheet1, Level2,scene):
    for i,j in enumerate(scene):
        sheet1['E'+str(i+3)] = calculate_FRR(int(Level2), j)
def learn_insert_L3(sheet1, Level3,scene):
    for i,j in enumerate(scene):
        sheet1['F'+str(i+3)] = calculate_FRR(int(Level3), j)
def learn_insert_L4(sheet1, Level4,scene):
    for i,j in enumerate(scene):
        sheet1['G'+str(i+3)] = calculate_FRR(int(Level4), j)
def no_insert_scene(sheet1,scene):
    for i,j in enumerate(scene):
        sheet1['C'+str(i+num+4)] = j
def no_insert_L1(sheet1, Level1,scene):
    for i,j in enumerate(scene):
        sheet1['D'+str(i+num+4)] = calculate_FRR(int(Level1), j)
def no_insert_L2(sheet1, Level2,scene):
    for i,j in enumerate(scene):
        sheet1['E'+str(i+num+4)] = calculate_FRR(int(Level2), j)
def no_insert_L3(sheet1,Level3,scene):
    for i,j in enumerate(scene):
        sheet1['F'+str(i+num+4)] = calculate_FRR(int(Level3), j)
def no_insert_L4(sheet1, Level4,scene):
    for i,j in enumerate(scene):
        sheet1['G'+str(i+num+4)] = calculate_FRR(int(Level4), j)

if __name__ == '__main__':
    scene=get_scene()
    num = int(len(scene))
    book = openpyxl.Workbook()
    sheet1 = book.active
    sheet1.title = 'FRR'
    sheet1["A1"] = '跑库信息'
    sheet1["B1"] = '学习'
    sheet1["C1"] = '场景'
    sheet1["D1"] = 'FAR'+config()[0]
    sheet1["E1"] = 'FAR'+config()[1]
    sheet1["F1"] = 'FAR'+config()[2]
    sheet1["G1"] = 'FAR'+config()[3]
    sheet1.merge_cells('A2:A'+str((num+1)*2+1))
    sheet1["A2"] = '预处理版本：'+config()[4]+' '+'match库：'+config()[5]+' '+'图像大小：'+config()[6]+' '+'跑库人数：'+config()[7]
    sheet1.merge_cells('B2:B'+str(num+2))
    sheet1.merge_cells('B'+str(num+3)+':B'+str((num+1)*2+1))
    sheet1["C2"] = "阈值"
    sheet1["C"+str(num+3)] = "阈值"
    Level1 = search_THS(config()[0])
    Level2 = search_THS(config()[1])
    Level3 = search_THS(config()[2])
    Level4 = search_THS(config()[3])
    if Learn_config()[0]== '1':
        sheet1["B2"] = 'Learn'+' '+'LT:'+Learn_config()[1]+' '+'LVT:'+Learn_config()[2]
        sheet1['D2'] = Level1
        sheet1['E2'] = Level2
        sheet1['F2'] = Level3
        sheet1['G2'] = Level4
        t1 = threading.Thread(target=learn_insert_scene,args=(sheet1,scene))
        t2 = threading.Thread(target=learn_insert_L1,args=(sheet1,Level1,scene))
        t3 = threading.Thread(target=learn_insert_L2,args=(sheet1,Level2,scene))
        t4 = threading.Thread(target=learn_insert_L3,args=(sheet1,Level3,scene))
        t5 = threading.Thread(target=learn_insert_L4,args=(sheet1,Level4,scene))
    else:
        sheet1['B'+str(num+3)] = 'nolearn'
        sheet1['D'+str(num+3)] = Level1
        sheet1['E'+str(num+3)] = Level2
        sheet1['F'+str(num+3)] = Level3
        sheet1['G'+str(num+3)] = Level4
        t1 = threading.Thread(target=no_insert_scene,args=(sheet1,scene))
        t2 = threading.Thread(target=no_insert_L1,args=(sheet1,Level1,scene))
        t3 = threading.Thread(target=no_insert_L2,args=(sheet1,Level2,scene))
        t4 = threading.Thread(target=no_insert_L3,args=(sheet1,Level3,scene))
        t5 = threading.Thread(target=no_insert_L4,args=(sheet1,Level4,scene))
    print('wait for calculate')
    t1.start()
    t2.start()
    t3.start()
    t4.start()
    t5.start()
    #判断线程是否全部跑完
    while True:
        if t1.isAlive()== False and t2.isAlive()== False and t3.isAlive()== False and t4.isAlive()== False and t5.isAlive()== False:
            thin = Side(border_style="thin")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            for row in sheet1['A1:G' + str((num + 1) * 2 + 1)]:
                for cell in row:
                    cell.border = border
                    cell.alignment = alignment
            book.save('result.xlsx')
            book.close()
            print('"Mission Complete"')
            break
        else:
            continue
